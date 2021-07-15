kahoot_18 = [
    r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\2018_kahoot\kahoot_12_09_2018.xlsx',
    r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\2018_kahoot\kahoot_12_09_2018_processos_design_ihc.xlsx',
    r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\2018_kahoot\kahoot_02_10_2018.xlsx',
    r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\2018_kahoot\kahoot_24_10_2018.xlsx',
    r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\2018_kahoot\kahoot_25_10_2018.xlsx',
    r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\2018_kahoot\kahoot_26_10_2018.xlsx',
    r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\2018_kahoot\kahoot_30_10_2018.xlsx',
    r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\2018_kahoot\kahoot_31_10_2018.xlsx',
    r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\2018_kahoot\kahoot_13_11_2018.xlsx',
    r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\2018_kahoot\kahoot_14_11_2018.xlsx']
kahoot_19 = [
    r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\2019_kahoot\results_04_09_2019_conceitos.xlsx',
    r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\2019_kahoot\results_04_09_2019_processos_design.xlsx',
    r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\2019_kahoot\results_01_10_2019_parte_1.xlsx',
    r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\2019_kahoot\results_01_10_2019_parte_2.xlsx',
    r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\2019_kahoot\resultado_09_10_2019.xlsx',
    r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\2019_kahoot\resultado_22_10_2019.xlsx',
    r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\2019_kahoot\resultado_29_10_2019.xlsx',
    r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\2019_kahoot\resultado_30_10_2019.xlsx',
    r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\2019_kahoot\resultado_05_11_2019.xlsx',
    r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\2019_kahoot\resultado_06_11_2019.xlsx',
    r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\2019_kahoot\resultado_19_11_2019.xlsx']

mapping_18 = r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\2018_kahoot\mapeamentos-e-notas-finais.xlsx'
mapping_19 = r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\2019_kahoot\notas_com_gamificacao.xlsx'

from openpyxl import load_workbook


class Lesson:
    def __init__(self):
        self.description = ""
        self.datetime = 0
        self.questions = []
        self.students = []
        self.scores = []


class Student(Lesson):
    def __init__(self):
        self.alias = ""


class Score(Lesson):
    def __init__(self, student, question, answer, isCorrect, isAttend):
        self.student = student
        self.question = question
        self.givenAnswer = answer
        self.isCorrect = isCorrect
        self.isAttend = isAttend

    def get_isCorrect(self):
        return self.isCorrect()

    def set_isCorrect(self, booleano):
        self.isCorrect = booleano

    def get_isAttend(self):
        return False

    def set_isAttend(self, booleano):
        self.isAttend = booleano


class Question(Lesson):
    def __init__(self):
        self.statement = ""
        self.option = []
        self.correctAnswer = 0


class Student_Answer:
    def __init__(self):
        self.name = ""
        self.arrayAnswer = []


def mappingStudent(file):
    workmapping = load_workbook(file)
    mapActive = workmapping.active

    namesSheets = workmapping.sheetnames

    mapping = workmapping[namesSheets[0]]
    nameStudents = []

    for students in range(2, mapActive.max_row+1):
        nameStudents.append(mapping.cell(row=students, column=1).value)

    return nameStudents


def vetor_zeros(quantidade):
    vetor = []
    for zeros in range(0, quantidade):
        vetor.append(0)

    return vetor


class Parser:
    def __init__(self):
        self.overview = {}
        self.finalScore = {}
        self.questionSummary = {}
        self.questions = {}
        self.nOfSheets = 0
        self.nOfQuestion = 0
        self.qtdAlunos = 0
        self.maxRow_QS = 0
        self.questionsList = 0
        self.question_list = []
        self.arrayAnswers = []
        self.arrayStudents = []

    def create_dict(self, file, year):
        workDict = load_workbook(file)
        dictActive = workDict.active

        self.nOfSheets = len(workDict.sheetnames)
        if year == 2018:
            self.nOfQuestion = self.nOfSheets - 4
        else:
            self.nOfQuestion = self.nOfSheets - 3
        names = workDict.sheetnames

        workOverview = workDict[names[0]]
        workFinalScore = workDict[names[1]]
        workQuestionSummary = workDict[names[2]]
        workQuestions = []

        for a in range(3, self.nOfQuestion+3):
            workQuestions.append(workDict[names[a]])

        self.overview = {
            'Basic Information': {
                "Lesson Name": workOverview.cell(row=1, column=1).value,
                "Played on": workOverview.cell(row=2, column=2).value,
                "Hosted by": workOverview.cell(row=3, column=2).value,
                "Played with": workOverview.cell(row=4, column=2).value,
                "Played": workOverview.cell(row=5, column=2).value
            },
            "Overall Performance": {
                "Total correct answers (%)": workOverview.cell(row=8, column=3).value,
                "Total incorrect answers (%)": workOverview.cell(row=9, column=3).value,
                "Average score (points)": workOverview.cell(row=10, column=3).value
            },
            "Feedback": {
                "How fun was it? (out of 5)": workOverview.cell(row=13, column=3).value,
                "Did you learn something?": {
                    "Yes": workOverview.cell(row=14, column=3).value,
                    "No": workOverview.cell(row=14, column=5).value
                },
                "Do you recommend it?": {
                    "Yes": workOverview.cell(row=15, column=3).value,
                    "No": workOverview.cell(row=15, column=5).value
                },
                "How do you feel?": {
                    "Positive": workOverview.cell(row=16, column=4).value,
                    "Neutral": workOverview.cell(row=16, column=6).value,
                    "Negative": workOverview.cell(row=16, column=8).value
                }
            }
        }

        # label_1 = wsOver.cell(row=7, column=1).value
        # label_2 = wsOver.cell(row=12, column=1).value

        if year == 2018:
            self.qtdAlunos = workFinalScore.max_row - 5
        else:
            self.qtdAlunos = workFinalScore.max_row - 3

        rank = []
        player = []
        total_score = []
        correct_answer = []
        incorrect_answer = []

        for b in range(4, self.qtdAlunos+4):
            rank.append(workFinalScore.cell(row=b, column=1).value)
            player.append(workFinalScore.cell(row=b, column=2).value)
            total_score.append(workFinalScore.cell(row=b, column=3).value)
            correct_answer.append(workFinalScore.cell(row=b, column=4).value)
            incorrect_answer.append(workFinalScore.cell(row=b, column=5).value)

        self.finalScore = {
            'Final Scores': {
                'rank': rank,
                'player': player,
                'totalScore': total_score,
                'correctAnswers': correct_answer,
                'incorrectAnswers': incorrect_answer
            }
        }

        qs_rank = []
        qs_player = []
        qs_totalScore = []

        qs_statement_question = []
        qs_questionStatment = []

        qs_score_question = []
        qs_questionsScore = []

        qs_answer_question = []
        qs_questionsAnswer = []

        for c in range(4, self.qtdAlunos+4):
            qs_rank.append(workQuestionSummary.cell(row=c, column=1).value)
            qs_player.append(workQuestionSummary.cell(row=c, column=2).value)
            qs_totalScore.append(workQuestionSummary.cell(row=c, column=3).value)
            for d in range(0, self.nOfQuestion):
                col_questions = 4 + 2*d
                col_statement = 5 + 2*d
                qs_statement_question.append(workQuestionSummary.cell(row=3, column=col_statement).value)
                qs_score_question.append(workQuestionSummary.cell(row=c, column=col_questions).value)
                qs_answer_question.append(workQuestionSummary.cell(row=c, column=col_statement).value)
            qs_questionStatment.append(qs_statement_question)
            qs_questionsScore.append(qs_score_question)
            qs_questionsAnswer.append(qs_answer_question)
            qs_statement_question = []
            qs_score_question = []
            qs_answer_question = []

        self.questionSummary = {'Question Summary': {
            'rank': qs_rank,
            'player': qs_player,
            'totalScore': qs_totalScore
        },
            1: {
                'score': qs_questionsScore,
                'statement': qs_questionStatment,
                'answer': qs_questionsAnswer
            }
        }

        question = []

        question_num = []
        question_statement = []
        question_correctAnswers = []
        question_playersCorrect = []
        question_questionDuration = []
        question_ansOptTriangle = []
        question_ansOptLosangle = []
        question_ansOptCircle = []
        question_ansOptSquare = []
        question_IsAnswerCorrectTriangle = []
        question_IsAnswerCorrectLosangle = []
        question_IsAnswerCorrectCircle = []
        question_IsAnswerCorrectSquare = []
        question_NumOfAnsReceivedTriangle = []
        question_NumOfAnsReceivedLosangle = []
        question_NumOfAnsReceivedCircle = []
        question_NumOfAnsReceivedSquare = []
        question_TimeToAnsTriangle = []
        question_TimeToAnsLosangle = []
        question_TimeToAnsCircle = []
        question_TimeToAnsSquare = []

        for questao in range(0, self.nOfQuestion):
            studentes_question = []
            questions = []

            for e in range(0, self.qtdAlunos):
                for f in range(1, 11):
                    studentes_question.append(workQuestions[questao].cell(row=e+15, column=f).value)
                questions.append(studentes_question)
                studentes_question = []
            question.append(questions)
            questions = []

            question_num.append(workQuestions[questao].cell(row=2, column=1).value)
            question_statement.append(workQuestions[questao].cell(row=2, column=2).value)
            question_correctAnswers.append(workQuestions[questao].cell(row=3, column=2).value)
            question_playersCorrect.append(workQuestions[questao].cell(row=4, column=3).value)
            question_questionDuration.append(workQuestions[questao].cell(row=5, column=3).value)
            question_ansOptTriangle.append(workQuestions[questao].cell(row=8, column=4).value)
            question_ansOptLosangle.append(workQuestions[questao].cell(row=8, column=6).value)
            question_ansOptCircle.append(workQuestions[questao].cell(row=8, column=8).value)
            question_ansOptSquare.append(workQuestions[questao].cell(row=8, column=10).value)

            if (workQuestions[questao].cell(row=9, column=3).value) == "✔︎":
                question_IsAnswerCorrectTriangle.append(True)
            else:
                question_IsAnswerCorrectTriangle.append(False)

            if (workQuestions[questao].cell(row=9, column=5).value) == "✔︎":
                question_IsAnswerCorrectLosangle.append(True)
            else:
                question_IsAnswerCorrectLosangle.append(False)

            if (workQuestions[questao].cell(row=9, column=7).value) == "✔︎":
                question_IsAnswerCorrectCircle.append(True)
            else:
                question_IsAnswerCorrectCircle.append(False)

            if (workQuestions[questao].cell(row=9, column=9).value) == "✔︎":
                question_IsAnswerCorrectSquare.append(True)
            else:
                question_IsAnswerCorrectSquare.append(False)

            question_NumOfAnsReceivedTriangle.append(workQuestions[questao].cell(row=10, column=3).value)
            question_NumOfAnsReceivedLosangle.append(workQuestions[questao].cell(row=10, column=5).value)
            question_NumOfAnsReceivedCircle.append(workQuestions[questao].cell(row=10, column=7).value)
            question_NumOfAnsReceivedSquare.append(workQuestions[questao].cell(row=10, column=9).value)
            question_TimeToAnsTriangle.append(workQuestions[questao].cell(row=11, column=3).value)
            question_TimeToAnsLosangle.append(workQuestions[questao].cell(row=11, column=5).value)
            question_TimeToAnsCircle.append(workQuestions[questao].cell(row=11, column=7).value)
            question_TimeToAnsSquare.append(workQuestions[questao].cell(row=11, column=9).value)

        #question[questao][aluno][informaão]

            if question[questao][0][2] == "✘":
                question[questao][0][2] = False
            else:
                question[questao][0][2] = True

            self.questions = {'Basic informations': {
                 'numOfQuestion': question_num[questao],
                 'statement': question_statement[questao],
                 'correctAnswers': question_correctAnswers[questao],
                 'playersCorrect': question_playersCorrect[questao],
                 'questionDuration': question_questionDuration[questao]
                },
                 "Answare Summary": {
                     'ansOptTriangle': question_ansOptTriangle[questao],
                     'ansOptLosangle': question_ansOptLosangle[questao],
                     'ansOptCircle': question_ansOptCircle[questao],
                     'ansOptSquare': question_TimeToAnsSquare[questao],
                     'IsAnswerCorrectTriangle': question_IsAnswerCorrectTriangle[questao],
                     'IsAnswerCorrectLosangle': question_IsAnswerCorrectLosangle[questao],
                     'IsAnswerCorrectCircle': question_IsAnswerCorrectCircle[questao],
                     'IsAnswerCorrectSquare': question_IsAnswerCorrectSquare[questao],
                     'NumOfAnsReceivedTriangle': question_NumOfAnsReceivedTriangle[questao],
                     'NumOfAnsReceivedLosangle': question_NumOfAnsReceivedLosangle[questao],
                     'NumOfAnsReceivedCircle': question_NumOfAnsReceivedCircle[questao],
                     'NumOfAnsReceivedSquare': question_NumOfAnsReceivedSquare[questao],
                     'TimeToAnsTriangle': question_TimeToAnsTriangle[questao],
                     'TimeToAnsLosangle': question_TimeToAnsLosangle[questao],
                     'TimeToAnsCircle': question_TimeToAnsCircle[questao],
                     'TimeToAnsSquare': question_TimeToAnsSquare[questao],
                    },
                 'Answer Details': {
                     1: {
                         'player': question[questao][0][0],
                         'alias': question[questao][0][1],
                         'answerIsCorrect': question[questao][0][2],
                         'statement': question[questao][0][3],
                         'score': question[questao][0][4],
                         'acumulateScore': question[questao][0][6],
                         'answerTime': question[questao][0][8]
                     }
                 }
            }
            for h in range(2, self.qtdAlunos+1):
                self.questions['Answer Details'][h] = {}
                _h = h - 1
                self.questions['Answer Details'][h]['player'] = question[questao][_h][0]
                self.questions['Answer Details'][h]['alias'] = question[questao][_h][1]
                if question[questao][_h][2] == "✘":
                    question[questao][_h][2] = False
                else:
                    question[questao][_h][2] = True
                self.questions['Answer Details'][h]['answerIsCorrect'] = question[questao][_h][2]
                self.questions['Answer Details'][h]['statement'] = question[questao][_h][3]
                self.questions['Answer Details'][h]['score'] = question[questao][_h][4]
                self.questions['Answer Details'][h]['acumulateScore'] = question[questao][_h][6]
                self.questions['Answer Details'][h]['answerTime'] = question[questao][_h][8]
            self.question_list.append(self.questions)

        #print(self.question_list[0]['Answer Details'][2])
        #print(self.questions['Answer Details'][2])
        #print(self.questions['Basic informations']['numOfQuestion'])

        #question[questao][aluno][informaão]
        #question_list[questao][chave][chave][atributo]

        return workDict

    def to_json(self, workbook):
        names = workbook.sheetnames

        txt_overview = ('{"' + names[0] + '": {' + "\n"
                        + "\t" + '"' + self.overview['Basic Information']["Lesson Name"] + '": {' + "\n"
                        + "\t\t" + '"' + "Played on" + '": "' + str(
                    self.overview['Basic Information']["Played on"]) + '", ' + "\n"
                        + "\t\t" + '"' + "Hosted by" + '": "' + str(
                    self.overview['Basic Information']["Hosted by"]) + '", ' + "\n"
                        + "\t\t" + '"' + "Played with" + '": "' + str(
                    self.overview['Basic Information']["Played with"]) + '", ' + "\n"
                        + "\t\t" + '"' + "Played" + '": "' + str(
                    self.overview['Basic Information']["Played"]) + '"' + "\n"
                        + "\t" + "}, " + "\n"
                        + "\t" + '"' + "Overall Performance" + '": {' + "\n"
                        + "\t\t" + '"' + "Total correct answers (%)" + '": "' + str(
                    self.overview['Overall Performance']["Total correct answers (%)"]) + '", ' + "\n"
                        + "\t\t" + '"' + "Total incorrect answers (%)" + '": "' + str(
                    self.overview['Overall Performance']["Total incorrect answers (%)"]) + '", ' + "\n"
                        + "\t\t" + '"' + "Average score (points)" + '": "' + str(
                    self.overview['Overall Performance']["Average score (points)"]) + '"' + "\n"
                        + "\t" + "}," + "\n"
                        + "\t" + '"' + "Feedback" + '": {' + "\n"
                        + "\t\t" + '"' + "How fun was it? (out of 5)" + '": "' + str(
                    self.overview['Feedback']["How fun was it? (out of 5)"]) + '", ' + "\n"
                        + "\t\t" + '"' + "Did you learn something?" + '": {' + "\n"
                        + "\t\t\t" + '"' + "Yes" + '": "' + str(
                    self.overview['Feedback']["Did you learn something?"]["Yes"]) + '", ' + "\n"
                        + "\t\t\t" + '"' + "No" + '": "' + str(
                    self.overview['Feedback']["Did you learn something?"]["No"]) + '"' + "\n"
                        + "\t\t" + "}," + "\n"
                        + "\t\t" + '"' + "Do you recommend it?" + '": {' + "\n"
                        + "\t\t\t" + '"' + "Yes" + '": "' + str(
                    self.overview['Feedback']["Do you recommend it?"]["Yes"]) + '", ' + "\n"
                        + "\t\t\t" + '"' + "No" + '": "' + str(
                    self.overview['Feedback']["Do you recommend it?"]["No"]) + '"' + "\n"
                        + "\t\t" + "}," + "\n"
                        + "\t\t" + '"' + "How do you feel?" + '": {' + "\n"
                        + "\t\t\t" + '"' + "Positive" + '": "' + str(
                    self.overview['Feedback']["How do you feel?"]["Positive"]) + '", ' + "\n"
                        + "\t\t\t" + '"' + "Neutral" + '": "' + str(
                    self.overview['Feedback']["How do you feel?"]["Neutral"]) + '", ' + "\n"
                        + "\t\t\t" + '"' + "Negative" + '": "' + str(
                    self.overview['Feedback']["How do you feel?"]["Negative"]) + '"' "\n"
                        + "\t\t" + "}" + "\n"
                        + "\t" + "}\n" + "}, \n")


        txt_final = '"' + names[1] + '": {' + "\n" + "\t" + '"' + self.overview['Basic Information'][
            "Lesson Name"] + '": \n' + "\t" + "["
        txt_final_student = []

        for i in range(0, self.qtdAlunos):
            txt_final_student.append("\t\t{\n"
                                     + "\t\t\t" + '"' + "Rank" + '": "' + str(
                self.finalScore['Final Scores']['rank'][i]) + '", ' + "\n"
                                     + "\t\t\t" + '"' + 'Player' + '": "' + str(
                self.finalScore['Final Scores']['player'][i]) + '", ' + "\n"
                                     + "\t\t\t" + '"' + 'Total Score' + '": "' + str(
                self.finalScore['Final Scores']['totalScore'][i]) + '", ' + "\n"
                                     + "\t\t\t" + '"' + 'Correct Answers' + '": "' + str(
                self.finalScore['Final Scores']['correctAnswers'][i]) + '", ' + "\n"
                                     + "\t\t\t" + '"' + 'Incorrect Answers' + '": "' + str(
                self.finalScore['Final Scores']['incorrectAnswers'][i]) + '"' + "\n")
            if i == self.qtdAlunos-1:
                txt_final_student[i] = txt_final_student[i] + "\t\t" + "}" + "\n"
            else:
                txt_final_student[i] = txt_final_student[i] + "\t\t" + "}," + "\n"
        txt_final_student[self.qtdAlunos-1] = txt_final_student[self.qtdAlunos-1] + "\t]\n},"


        txt_questionSummary_students = []
        txt_questionSummary_students_questions = []
        concate = []

        txt_questionSummary = '"' + names[2] + '": {' + "\n" + "\t" + '"' + self.overview['Basic Information'][
            "Lesson Name"] + '": {' \
                              + "\n" + "\t\t" + '"' + names[2] + '": ' + "\n" + "\t\t" + "["
        n = 1

        # questionSummary["Alunos"]['score'][0][1]
        # questionSummary["Alunos"][_CAMPO_][_QUESTÃO_][_ALUNO_]

        for j in range(0, self.qtdAlunos):
            txt_questionSummary_students.append("\t\t\t" + "{\n"
                                                + "\t\t\t\t" + '"' + 'Rank' + '": "' + str(
                self.questionSummary['Question Summary']['rank'][j]) + '", ' + "\n"
                                                + "\t\t\t\t" + '"' + 'Player' + '": "' + str(
                self.questionSummary['Question Summary']['player'][j]) + '", ' + "\n"
                                                + "\t\t\t\t" + '"' + 'Total Score' + '": "' + str(
                self.questionSummary['Question Summary']['totalScore'][j]) + '", ' + "\n")
            concatenar = txt_questionSummary_students[j]
            for k in range(0, self.nOfQuestion):
                txt_questionSummary_students_questions.append("\t\t\t\t" + '"' + 'Q' + str(k + 1) + '": {' + "\n"
                                                              + "\t\t\t\t\t\t" + '"' + "Score" + '": "' + str(
                    self.questionSummary[1]['score'][j][k]) + '", ' + "\n"
                                                              + "\t\t\t\t\t\t" + '"' + "Statement" + '": "' + str(
                    self.questionSummary[1]['statement'][j][k]) + '", ' + "\n"
                                                              + "\t\t\t\t\t\t" + '"' + "Answer" + '": "' + str(
                    self.questionSummary[1]['answer'][j][k]) + '"' + "\n")
                if k == self.nOfQuestion - 1:
                    txt_questionSummary_students_questions[k] = txt_questionSummary_students_questions[k] + (
                            "\t\t\t\t" + "}" + "\n")
                else:
                    txt_questionSummary_students_questions[k] = txt_questionSummary_students_questions[
                                                                    k] + "\t\t\t\t" + "}, " + "\n"
                concatenar = concatenar + txt_questionSummary_students_questions[k]
            txt_questionSummary_students_questions = []
            if j == self.qtdAlunos - 1:
                concatenar = concatenar + "\t\t\t" + "}" + "\n"
            else:
                concatenar = concatenar + "\t\t\t" + "}, " + "\n"
            concate.append(concatenar)

        ind = self.qtdAlunos - 1
        concate[ind] = concate[ind] + "\t\t]\n" + "\t" + "}\n" + "},\n"

        txt_question = []
        txt_question_alunos = []
        concate_question = []

        for questao in range(0, self.nOfQuestion):
            txt_question.append('"' + names[questao + 3] + '": {' + "\n" + "\t"
                                + '"' + self.overview['Basic Information']["Lesson Name"] + '": {' + "\n"
                                + "\t\t" + '"' + self.question_list[questao]['Basic informations']['numOfQuestion']
                                + '": {' + "\n"
                                + "\t\t\t" + '"' + "Statement" + '": "' +
                                self.question_list[questao]['Basic informations']['statement']
                                + '", ' + "\n"
                                + "\t\t\t" + '"' + "Correct Answer" + '": "' + str(
                self.question_list[questao]['Basic informations']['correctAnswers']) + '", ' + "\n"
                                + "\t\t\t" + '"' + "Players Correct" + '": "' + str(
                self.question_list[questao]['Basic informations']['playersCorrect']) + '", ' + "\n"
                                + "\t\t\t" + '"' + "Question Duration" + '": "' + str(
                self.question_list[questao]['Basic informations']['questionDuration']) + '"' + "\n"
                                + "\t\t" + "},\n"
                                + "\t\t" + '"' + "Answare Summary" + '": {' + "\n"
                                + "\t\t\t" + '"' + "Answer options" + '": {' + "\n"
                                + "\t\t\t\t\t" + '"' + "Triangle" + '": ' + str(
                self.question_list[questao]['Answare Summary']['ansOptTriangle']) + ', ' + "\n"
                                + "\t\t\t\t\t" + '"' + "Losangle" + '": ' + str(
                self.question_list[questao]['Answare Summary']['ansOptLosangle']) + ', ' + "\n"
                                + "\t\t\t\t\t" + '"' + "Circle" + '": ' + str(
                self.question_list[questao]['Answare Summary']['ansOptCircle']) + ', ' + "\n"
                                + "\t\t\t\t\t" + '"' + "Square" + '": ' + str(
                self.question_list[questao]['Answare Summary']['ansOptSquare']) + '' + "\n"
                                + "\t\t\t" + "},\n"
                                + "\t\t\t" + '"' + "Is answer correct?" + '": {' + "\n"
                                + "\t\t\t\t\t" + '"' + "Triangle" + '": "' + str(
                self.question_list[questao]['Answare Summary']['IsAnswerCorrectTriangle']) + '", ' + "\n"
                                + "\t\t\t\t\t" + '"' + "Losangle" + '": "' + str(
                self.question_list[questao]['Answare Summary']['IsAnswerCorrectLosangle']) + '", ' + "\n"
                                + "\t\t\t\t\t" + '"' + "Circle" + '": "' + str(
                self.question_list[questao]['Answare Summary']['IsAnswerCorrectCircle']) + '", ' + "\n"
                                + "\t\t\t\t\t" + '"' + "Square" + '": "' + str(
                self.question_list[questao]['Answare Summary']['IsAnswerCorrectSquare']) + '"' + "\n"
                                + "\t\t\t" + "},\n"
                                + "\t\t\t" + '"' + "Number of answers received" + '": {' + "\n"
                                + "\t\t\t\t\t" + '"' + "Triangle" + '": "' + str(
                self.question_list[questao]['Answare Summary']['NumOfAnsReceivedTriangle']) + '", ' + "\n"
                                + "\t\t\t\t\t" + '"' + "Losangle" + '": "' + str(
                self.question_list[questao]['Answare Summary']['NumOfAnsReceivedLosangle']) + '", ' + "\n"
                                + "\t\t\t\t\t" + '"' + "Circle" + '": "' + str(
                self.question_list[questao]['Answare Summary']['NumOfAnsReceivedCircle']) + '", ' + "\n"
                                + "\t\t\t\t\t" + '"' + "Square" + '": "' + str(
                self.question_list[questao]['Answare Summary']['NumOfAnsReceivedSquare']) + '"' + "\n"
                                + "\t\t\t" + "},\n"
                                + "\t\t\t" + '"' + "Average time taken to answer (seconds)" + '": {' + "\n"
                                + "\t\t\t\t\t" + '"' + "Triangle" + '": "' + str(
                self.question_list[questao]['Answare Summary']['TimeToAnsTriangle']) + '", ' + "\n"
                                + "\t\t\t\t\t" + '"' + "Losangle" + '": "' + str(
                self.question_list[questao]['Answare Summary']['TimeToAnsLosangle']) + '", ' + "\n"
                                + "\t\t\t\t\t" + '"' + "Circle" + '": "' + str(
                self.question_list[questao]['Answare Summary']['TimeToAnsCircle']) + '", ' + "\n"
                                + "\t\t\t\t\t" + '"' + "Square" + '": "' + str(
                self.question_list[questao]['Answare Summary']['TimeToAnsSquare']) + '"' + "\n"
                                + "\t\t\t" + "}\n"
                                + "\t\t" + "},\n"
                                + "\t\t" + '"' + "Answer Details" + '": ' + "\n" + "\t\t[\n")
            concatenar_Question = txt_question[questao]
            for alunos in range(0, self.qtdAlunos):
                aluno = alunos + 1
                txt_question_alunos.append("\t\t\t" + "{\n"
                                           + "\t\t\t\t" + '"' + "Player" + '": "' + str(
                    self.question_list[questao]['Answer Details'][aluno]['player']) + '", ' + "\n"
                                           + "\t\t\t\t" + '"' + "Alias" + '": "' + str(
                    self.question_list[questao]['Answer Details'][aluno]['alias']) + '", ' + "\n"
                                           + "\t\t\t\t" + '"' + "Answer" + '": {' + "\n"
                                           + "\t\t\t\t\t\t" + '"' + "Correct?" + '": "' + str(
                    self.question_list[questao]['Answer Details'][aluno]['answerIsCorrect']) + '", ' + "\n"
                                           + "\t\t\t\t\t\t" + '"' + "Statement" + '": "' + str(
                    self.question_list[questao]['Answer Details'][aluno]['statement']) + '" ' + "\n"
                                           + "\t\t\t\t" + "}," + "\n"
                                           + "\t\t\t\t" + '"' + "Score (points)" + '": "' + str(
                    self.question_list[questao]['Answer Details'][aluno]['score']) + '", ' + "\n"
                                           + "\t\t\t\t" + '"' + "Current Total Score (points)" + '": "' + str(
                    self.question_list[questao]['Answer Details'][aluno]['acumulateScore']) + '", ' + "\n"
                                           + "\t\t\t\t" + '"' + "Answer time (seconds)" + '": "' + str(
                    self.question_list[questao]['Answer Details'][aluno]['answerTime']) + '"' + "\n")
                if alunos == self.qtdAlunos - 1:
                    txt_question_alunos[alunos] = txt_question_alunos[alunos] + "\t\t\t" + "}\n"
                else:
                    txt_question_alunos[alunos] = txt_question_alunos[alunos] + "\t\t\t" + "},\n"
                concatenar_Question = concatenar_Question + txt_question_alunos[alunos]
            txt_question_alunos = []
            concatenar_Question = concatenar_Question + "\t\t" + "]\n" + "\t" + "}\n"
            if questao == self.nOfQuestion - 1:
                concatenar_Question = concatenar_Question + "}\n}"
            else:
                concatenar_Question = concatenar_Question + "},\n"
            concate_question.append(concatenar_Question)

        arquivo = open('C:/Users/lucas/OneDrive/Documentos/UFG/9º Período - EXTRA/Projeto Final de Curso 2/data/novo_json/' + self.overview['Basic Information']["Lesson Name"] + ".json", 'w', encoding="utf-8")
        arquivo.write(txt_overview)
        arquivo.write("\n")
        arquivo.write(txt_final)
        arquivo.write(''.join(txt_final_student))
        arquivo.write("\n")
        arquivo.write(''.join(txt_questionSummary))
        arquivo.write(''.join(concate))
        arquivo.write("\n")
        arquivo.write(''.join(concate_question))
        arquivo.close()

    def to_lesson(self):
        this_lesson = Lesson()

        this_lesson.description = self.overview['Basic Information']["Lesson Name"]
        this_lesson.datetime = self.overview['Basic Information']["Played on"]

        students = []
        questions = []
        score = []

        for o in range(0, self.nOfQuestion):
            questions.append(Question())
            questions[o].statement = self.question_list[o]['Basic informations']['statement']
            questions[o].option.append(self.question_list[o]['Answare Summary']['ansOptTriangle'])
            questions[o].option.append(self.question_list[o]['Answare Summary']['ansOptLosangle'])
            questions[o].option.append(self.question_list[o]['Answare Summary']['ansOptCircle'])
            questions[o].option.append(self.question_list[o]['Answare Summary']['ansOptSquare'])
            if str(self.question_list[o]['Answare Summary']['IsAnswerCorrectTriangle']) == "True":
                questions[o].correctAnswer = 1
            if str(self.question_list[o]['Answare Summary']['IsAnswerCorrectLosangle']) == "True":
                questions[o].correctAnswer = 2
            if str(self.question_list[o]['Answare Summary']['IsAnswerCorrectCircle']) == "True":
                questions[o].correctAnswer = 3
            if str(self.question_list[o]['Answare Summary']['IsAnswerCorrectSquare']) == "True":
                questions[o].correctAnswer = 4

        for p in range(0, self.qtdAlunos):
            q = p + 1
            students.append(Student())
            students[p].alias = self.question_list[0]['Answer Details'][q]['alias']
            this_lesson.students.append(students[p])

        for r in range(0, self.qtdAlunos):
            for s in range(0, self.nOfQuestion):
                r_ = r + 1
                answer = self.question_list[s]['Answer Details'][r_]['statement']
                correct = self.question_list[s]['Answer Details'][r_]['answerIsCorrect']
                if answer == "" or answer == " ":
                    attend = False
                else:
                    attend = True
                score.append(Score(students[r], questions[s], answer, correct, attend))

        this_lesson.students = students
        this_lesson.questions = questions
        this_lesson.scores = score

        return this_lesson

    def to_array(self, lesson):
        n = self.nOfQuestion
        c = 0
        for t in range(0, len(lesson.scores)):
            c = c + 1
            if lesson.scores[t].isAttend == False:
                self.arrayAnswers.append(0)
            elif lesson.scores[t].isCorrect == True:
                self.arrayAnswers.append(1)
            elif lesson.scores[t].isCorrect == False:
                self.arrayAnswers.append(-1)
            if c == n:
                self.arrayStudents.append(self.arrayAnswers)
                c = 0
                self.arrayAnswers = []


def to_arff(array, named):
    arff = open('C:/Users/lucas/OneDrive/Documentos/UFG/9º Período - EXTRA/Projeto Final de Curso 2/data/Arff_2018/' + named + ".arff", 'w', encoding="utf-8")

    arff.write('@relation ' + named + "\n\n")

    for i in range(len(array[0])):
        arff.write('@attribute Questao_' + str(i+1))
        arff.write("{'1', '0', '-1'}\n")

    arff.write("@attribute situacao {'fortemente_aprovado', 'fortemente_reprovado'}\n")
    arff.write('\n\n\n')
    arff.write('@data\n')

    for aux1 in range(len(array)):
        for aux2 in range(len(array[aux1])):
            arff.write("'")
            arff.write(str(array[aux1][aux2]))
            arff.write("'")
            if aux2 != len(array[aux1])-1:
                arff.write(", ")
        arff.write("\n\n")
    arff.close()


def set_situacao(array):
    zero_count = 0
    um_count = 0
    menosum_count = 0
    for i in range(len(array)):
        if array[i] == 0:
            zero_count = zero_count + 1
        elif array[i] == 1:
            um_count = um_count + 1
        elif array[i] == -1:
            menosum_count = menosum_count + 1

    if um_count >= 0.6*(len(array)):
        situacao = "fortmente_aprovado"
    else:
        situacao = "fortemente_reprovado"
    return situacao
#================== TESTES ===================================


alunos18 = mappingStudent(mapping_18)

students18 = []
for aluno in range(0, len(alunos18)):
    students18.append(Student_Answer())
    students18[aluno].name = alunos18[aluno]

read18 = []
dict18 = []
lesson18 = []

for obj in range(0, len(kahoot_18)):
    read18.append(Parser())
    dict18.append(read18[obj].create_dict(kahoot_18[obj], 2018))
    read18[obj].to_json(dict18[obj])
    lesson18.append(read18[obj].to_lesson())
    read18[obj].to_array(lesson18[obj])

alunos_aula18 = []
faltantes_aula18 = []
aulas18 = []
frequencias18 = []

for lesson in lesson18:
    for aluno18 in lesson.students:
        alunos_aula18.append(aluno18.alias)
    aulas18.append(alunos_aula18)
    alunos_aula18 = []

#print(alunos18)
#print(aulas18[3])
for aula in range(0, len(aulas18)):
    for frequencia18 in range(0, len(alunos18)):
        if alunos18[frequencia18] not in aulas18[aula]:
            faltantes_aula18.append(alunos18[frequencia18])
    frequencias18.append(faltantes_aula18)
    faltantes_aula18 = []

#print(frequencias18[0])

alunos = {}
for nome in alunos18:
    alunos[nome] = []

dias = []
dia = {}
vetor_aula = []
for aulas in range(0, len(kahoot_18)):
    z = 0
    for day in aulas18[aulas]:
        dia["NumQ"] = len(read18[aulas].arrayStudents[z])
        dia[day] = []
        dia[day].append(read18[aulas].arrayStudents[z])
        z = z + 1
    dias.append(dia)
    #print(dias[aulas])
    dia = {}

for dia in dias:
    chaves = dia.keys()
    for aluno in chaves:
        if aluno == "NumQ":
            continue
        if aluno not in alunos:
            continue
        alunos[aluno].append(dia[aluno])
    alunosChaves = alunos.keys()
    for aluno in alunosChaves:
        if aluno == "NumQ":
            continue
        if aluno not in dia:
            alunos[aluno].append(vetor_zeros(dia["NumQ"]))

from pandas.core.common import flatten

divisao_15 = []
divisao_35 = []
divisao_50 = []
divisao_69 = []
alunos_chaves = alunos.keys()
for aluno in alunos_chaves:
    alunos[aluno] = list(flatten(alunos[aluno]))
    slice_15 = slice(15)
    slice_35 = slice(35)
    slice_50 = slice(50)
    slice_69 = slice(69)
    divisao_15.append(alunos[aluno][slice_15])
    divisao_35.append(alunos[aluno][slice_35])
    divisao_50.append(alunos[aluno][slice_50])
    divisao_69.append(alunos[aluno][slice_69])

for sit in range(len(divisao_15)):
    divisao_15[sit].append(set_situacao(divisao_15[sit]))

for sit in range(len(divisao_35)):
    divisao_35[sit].append(set_situacao(divisao_35[sit]))

for sit in range(len(divisao_50)):
    divisao_50[sit].append(set_situacao(divisao_50[sit]))

for sit in range(len(divisao_69)):
    divisao_69[sit].append(set_situacao(divisao_69[sit]))

to_arff(divisao_15, "Arff_15")
to_arff(divisao_35, "Arff_35")
to_arff(divisao_50, "Arff_50")
to_arff(divisao_69, "Arff_Total")


# c = Classifier(name='weka.classifiers.lazy.IBk', ckargs={'-K':1})
# c.train('training.arff')
# predictions = c.predict('query.arff')
#=======================================================================


#alunos19 = mappingStudent(mapping_19)
#students19 = []
# read2 = Parser()
# read2.createDict(kahoot_19[3], 2019)
# print(read2.questionSummary)

alunos19 = mappingStudent(mapping_19)

students19 = []
for aluno in range(0, len(alunos19)):
    students19.append(Student_Answer())
    students19[aluno].name = alunos19[aluno]

read19 = []
dict19 = []
lesson19 = []

for obj in range(0, len(kahoot_19)):
    read19.append(Parser())
    dict19.append(read19[obj].create_dict(kahoot_19[obj], 2019))
    read19[obj].to_json(dict19[obj])
    lesson19.append(read19[obj].to_lesson())
    read19[obj].to_array(lesson19[obj])

alunos_aula19 = []
faltantes_aula19 = []
aulas19 = []
frequencias19 = []

for lesson in lesson19:
    for aluno19 in lesson.students:
        alunos_aula19.append(aluno19.alias)
    aulas19.append(alunos_aula19)
    alunos_aula19 = []

#print(alunos19)
#print(aulas19[3])
for aula in range(0, len(aulas19)):
    for frequencia19 in range(0, len(alunos19)):
        if alunos19[frequencia19] not in aulas19[aula]:
            faltantes_aula19.append(alunos19[frequencia19])
    frequencias19.append(faltantes_aula19)
    faltantes_aula19 = []

#print(frequencias19[0])

alunos_19 = {}
for nome in alunos19:
    alunos_19[nome] = []

dias19 = []
dia19 = {}
vetor_aula19 = []
for aulas in range(0, len(kahoot_19)):
    z19 = 0
    for day in aulas19[aulas]:
        dia19["NumQ"] = len(read19[aulas].arrayStudents[z19])
        dia19[day] = []
        dia19[day].append(read19[aulas].arrayStudents[z19])
        z19 = z19 + 1
    dias19.append(dia19)
    dia19 = {}

for dia in dias19:
    chaves19 = dia.keys()
    for aluno in chaves19:
        if aluno == "NumQ":
            continue
        if aluno not in alunos_19:
            continue
        alunos_19[aluno].append(dia[aluno])

    alunosChaves19 = alunos_19.keys()
    for aluno in alunosChaves19:
        if aluno == "NumQ":
            continue
        if aluno not in dia:
            alunos_19[aluno].append(vetor_zeros(dia["NumQ"]))

from pandas.core.common import flatten

divisao_15_19 = []
divisao_35_19 = []
divisao_50_19 = []
alunos_chaves19 = alunos_19.keys()
for aluno in alunos_chaves19:
    alunos_19[aluno] = list(flatten(alunos_19[aluno]))
    slice_15_19 = slice(15)
    slice_35_19 = slice(35)
    slice_50_19 = slice(50)
    divisao_15_19.append(alunos_19[aluno][slice_15_19])
    divisao_35_19.append(alunos_19[aluno][slice_35_19])
    divisao_50_19.append(alunos_19[aluno][slice_50_19])
