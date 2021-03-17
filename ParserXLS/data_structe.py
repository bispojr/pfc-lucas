#Comentarios

#overview[1]['playedWith'] = '20'
#finalScore[1]['rank'] = 1
#finalScore[3] = {}
#finalScore[3]['player'] = "Animal"
#questions = { 1: {'score': "", 'statement': "", 'answer': ""}}
#questions[0] = d
#questions[2] = {}
#questions[2][1]['player'] = "Lucas"

#=================================================
#       Estruturas

#overview = { 1: { 'playedOn': "", 'hostedBy': "", 'playedWith': "", 'played': "", 'CorrectAnswers' : "", 'IncorrectAnswers': "", 'Average score': ""}}
#finalScore = { 1: { 'rank': "", 'player': "", 'totalScore': "", 'correctAnswers': "", 'incorrectAnswers': ""}}
#questionSummary = { 1: { 'rank': "", 'player': "", 'totalScore': "", 1: {'score': "", 'statement': "", 'answer': ""}}}
#questions = {'statement': "",'correctAnswers': "", 'playersCorrect': "",'questionDuration': "",'ansOptTriangle': "",'ansOptLosangle': "",'ansOptCircle': "",'ansOptSquare': "",'IsAnswerCorrectTriangle': "",'IsAnswerCorrectLosangle': "",'IsAnswerCorrectCircle': "",'IsAnswerCorrectSquare': "",'NumOfAnsReceivedTriangle': "", 'NumOfAnsReceivedLosangle': "",'NumOfAnsReceivedCircle': "",'NumOfAnsReceivedSquare': "",'TimeToAnsTriangle': "", 'TimeToAnsLosangle': "", 'TimeToAnsCircle': "", 'TimeToAnsSquare': "",
 #   1: {
  #      'player': "", 'alias': "", 'answerIsCorrect': "", 'statement': "", 'score': "", 'acumulateScore': "", 'answerTime': ""}}

#=============================================

from openpyxl import Workbook
from openpyxl import load_workbook
kahoot_1209 = r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\kahoot\2018\kahoot_12_09_2018.xlsx'
kahoot_1209_processo = r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\kahoot\2018\kahoot_12_09_2018_processos_design_ihc.xlsx'
kahoot_0210 = r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\kahoot\2018\kahoot_02_10_2018.xlsx'
kahoot_2410 = r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\kahoot\2018\kahoot_24_10_2018.xlsx'
kahoot_2510 = r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\kahoot\2018\kahoot_25_10_2018.xlsx'
kahoot_2610 = r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\kahoot\2018\kahoot_26_10_2018.xlsx'
kahoot_3010 = r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\kahoot\2018\kahoot_30_10_2018.xlsx'
kahoot_3110 = r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\kahoot\2018\kahoot_31_10_2018.xlsx'
kahoot_1311 = r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\kahoot\2018\kahoot_13_11_2018.xlsx'
kahoot_1411 = r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\kahoot\2018\kahoot_14_10_2018.xlsx'

wb = load_workbook(kahoot_0210)
ws = wb.active

nOfSheets = len(wb.worksheets)

wsOver = wb['Overview']
wsFinal = wb['Final Scores']
wsQS = wb['Question Summary']
wsQues = []

for q in range(0, (nOfSheets-3)):
    n = str(q+1)
    wsQues.append(wb['Question ' + n])

date = wsOver.cell(row=2, column=2).value
teacher = wsOver.cell(row=3, column=2).value
nOfPlayers = wsOver.cell(row=4, column=2).value
nOfQuestions = wsOver.cell(row=5, column=2).value
totalCorrect = wsOver.cell(row=8, column=3).value
totalIncorrect = wsOver.cell(row=9, column=3).value
averageScore = wsOver.cell(row=10, column=3).value

overview = { 1: { 
    'playedOn': date, 
    'hostedBy': teacher, 
    'playedWith': nOfPlayers, 
    'played': nOfQuestions, 
    'CorrectAnswers' : totalCorrect, 
    'IncorrectAnswers': totalIncorrect, 
    'Average score': averageScore
    }
}

#print(overview)
'''
print(date)
print(teacher)
print(nOfPlayers)
print(totalCorrect*100, "%")
print(totalIncorrect*100, "%")
print(averageScore)
'''

#Lista de numeros
idAl = []
#Lista de alunos
student = []
#Lista de score
score = []
#lista de questoes certas
quesCorrect = []
#lista de questoes erradas
quesIncorrect = []

stopCount = (int(nOfPlayers))+4
#print(stopCount)

#loop 1: 4 até nOfPlayers+3
    #col=1 and row = i 
for ial in range(4, stopCount):
    idAl.append(wsFinal.cell(row=ial, column=1).value)

#loop 2: 4 até nOfPlayers+3
    #col=2 and row = i
for istd in range(4, stopCount):
    student.append(wsFinal.cell(row=istd, column=2).value)

#loop 3: 4 até nOfPlayers+3
    #col=3 and row = i
for iscor in range(4, stopCount):
    score.append(wsFinal.cell(row=iscor, column=3).value)

#loop 4: 4 até nOfPlayers+3
    #col=4 and row = i
for iqc in range(4, stopCount):
    quesCorrect.append(wsFinal.cell(row=iqc, column=4).value)

#loop 5: 4 até nOfPlayers+3
    #col=5 and row = i
for iqi in range(4, stopCount):
    quesIncorrect.append(wsFinal.cell(row=iqi, column=5).value)

finalScore = { 1: { 
    'rank': idAl,
    'player': student, 
    'totalScore': score, 
    'correctAnswers': quesCorrect, 
    'incorrectAnswers': quesIncorrect
    }
}

#print(finalScore)

'''
for iprint in range(0, stopCount-4):
    print(idAl[iprint], "|", student[iprint], "|", score[iprint], "|", quesCorrect[iprint], "|", quesIncorrect[iprint])
'''

qs_idAl = []
qs_student = []
qs_totalScore = []

qs_count = (int(nOfQuestions))+1
qs_score = [[] for _ in range(qs_count)]
qs_statement = [[] for _s_ in range(qs_count)]
qs_answer = [[] for __ in range(qs_count)]

for i in range(4, stopCount):
    qs_idAl.append(wsQS.cell(row=i, column=1).value)
    qs_student.append(wsQS.cell(row=i, column=2).value)
    qs_totalScore.append(wsQS.cell(row=i, column=3).value)
    for j in range(0, qs_count):
        col = 4+2*j
        col2 = 5+2*j
        p = i-1
        qs_score[j].append(wsQS.cell(row=i, column=col).value)
        qs_statement[j].append(wsQS.cell(row=p, column=col2).value)
        qs_answer[j].append(wsQS.cell(row=i, column=col2).value)         

questionSummary = { 1: { 
    'rank': qs_idAl, 
    'player': qs_student, 
    'totalScore': qs_totalScore, 
                    1: {
    'score': qs_score, 
    'statement': qs_statement, 
    'answer': qs_answer
                        }
                    }
                }

#print(questionSummary)    
'''
for iprint in range(0, stopCount-4):
    print(qs_idAl[iprint], "|", qs_student[iprint], "|", qs_totalScore[iprint], "|")
    for qprint in range(0, qs_count-1):
        print(qs_score[qprint][iprint], "|", qs_answer[qprint][iprint], "|")
    print("")
'''

qtdAlunos = nOfPlayers+1
for quest in range(0, 5):

    wsq_statement = []
    wsq_correctAnswers = []
    wsq_playersCorrect = []
    wsq_questionDuration = []
    wsq_ansOptTriangle = []
    wsq_ansOptLosangle = []
    wsq_ansOptCircle = []
    wsq_ansOptSquare = []
    wsq_IsAnswerCorrectTriangle = []
    wsq_IsAnswerCorrectLosangle = []
    wsq_IsAnswerCorrectCircle = []
    wsq_IsAnswerCorrectSquare = []
    wsq_NumOfAnsReceivedTriangle = []
    wsq_NumOfAnsReceivedLosangle = []
    wsq_NumOfAnsReceivedCircle = []
    wsq_NumOfAnsReceivedSquare = []
    wsq_TimeToAnsTriangle = []
    wsq_TimeToAnsLosangle = []
    wsq_TimeToAnsCircle = []
    wsq_TimeToAnsSquare = []

    wsq_alunos = [[] for ____ in range(qtdAlunos)]

    for k in range(0, qtdAlunos):
        for l in range(1, 11):
            wsq_alunos[k].append(wsQues[quest].cell(row=k+15, column=l).value)

    wsq_statement.append(wsQues[quest].cell(row=2, column=2).value)
    wsq_correctAnswers.append(wsQues[quest].cell(row=3, column=3).value)
    wsq_playersCorrect.append(wsQues[quest].cell(row=4, column=3).value)
    wsq_questionDuration.append(wsQues[quest].cell(row=5, column=3).value)
    wsq_ansOptTriangle.append(wsQues[quest].cell(row=8, column=4).value)
    wsq_ansOptLosangle.append(wsQues[quest].cell(row=8, column=6).value)
    wsq_ansOptCircle.append(wsQues[quest].cell(row=8, column=8).value)
    wsq_ansOptSquare.append(wsQues[quest].cell(row=8, column=10).value)

    if (wsQues[quest].cell(row=9, column=3).value) == "✔︎":
        wsq_IsAnswerCorrectTriangle.append(True)
    else:
        wsq_IsAnswerCorrectTriangle.append(False)

    if (wsQues[quest].cell(row=9, column=5).value) == "✔︎":
        wsq_IsAnswerCorrectLosangle.append(True)
    else:
        wsq_IsAnswerCorrectLosangle.append(False)

    if (wsQues[quest].cell(row=9, column=7).value) == "✔︎":
        wsq_IsAnswerCorrectCircle.append(True)
    else:
        wsq_IsAnswerCorrectCircle.append(False)

    if (wsQues[quest].cell(row=9, column=9).value) == "✔︎":
        wsq_IsAnswerCorrectSquare.append(True)
    else:
        wsq_IsAnswerCorrectSquare.append(False)

    wsq_NumOfAnsReceivedTriangle.append(wsQues[quest].cell(row=10, column=3).value) 
    wsq_NumOfAnsReceivedLosangle.append(wsQues[quest].cell(row=10, column=5).value)
    wsq_NumOfAnsReceivedCircle.append(wsQues[quest].cell(row=10, column=7).value)
    wsq_NumOfAnsReceivedSquare.append(wsQues[quest].cell(row=10, column=9).value) 
    wsq_TimeToAnsTriangle.append(wsQues[quest].cell(row=11, column=3).value) 
    wsq_TimeToAnsLosangle.append(wsQues[quest].cell(row=11, column=5).value)
    wsq_TimeToAnsCircle.append(wsQues[quest].cell(row=11, column=7).value)
    wsq_TimeToAnsSquare.append(wsQues[quest].cell(row=11, column=9).value)

    questions = { 1: {
        'statement': wsq_statement,
        'correctAnswers': wsq_correctAnswers, 
        'playersCorrect': wsq_playersCorrect,
        'questionDuration': wsq_questionDuration,
        'ansOptTriangle': wsq_ansOptTriangle,
        'ansOptLosangle': wsq_ansOptLosangle,
        'ansOptCircle': wsq_ansOptCircle,
        'ansOptSquare': wsq_ansOptSquare,
        'IsAnswerCorrectTriangle': wsq_IsAnswerCorrectTriangle,  
        'IsAnswerCorrectLosangle': wsq_IsAnswerCorrectLosangle,
        'IsAnswerCorrectCircle': wsq_IsAnswerCorrectCircle,
        'IsAnswerCorrectSquare': wsq_IsAnswerCorrectSquare,
        'NumOfAnsReceivedTriangle': wsq_NumOfAnsReceivedTriangle,  
        'NumOfAnsReceivedLosangle': wsq_NumOfAnsReceivedLosangle,
        'NumOfAnsReceivedCircle': wsq_NumOfAnsReceivedCircle,
        'NumOfAnsReceivedSquare': wsq_NumOfAnsReceivedSquare,
        'TimeToAnsTriangle': wsq_TimeToAnsTriangle,  
        'TimeToAnsLosangle': wsq_TimeToAnsLosangle,
        'TimeToAnsCircle': wsq_TimeToAnsCircle,
        'TimeToAnsSquare': wsq_TimeToAnsSquare,
    1: {
        'player': wsq_alunos[0][0],
        'alias': wsq_alunos[0][1],
        'answerIsCorrect': wsq_alunos[0][2],
        'statement': wsq_alunos[0][3],
        'score': wsq_alunos[0][4],
        'acumulateScore': wsq_alunos[0][6],
        'answerTime': wsq_alunos[0][8]
    }
}}
    for alunos in range(2, nOfPlayers):
        questions[alunos] = {}
        
        questions[alunos]['player'] = wsq_alunos[alunos][0]
        questions[alunos]['alias'] = wsq_alunos[alunos][1]
        questions[alunos]['answerIsCorrect'] = wsq_alunos[alunos][2]
        questions[alunos]['statement'] = wsq_alunos[alunos][3]
        questions[alunos]['score'] = wsq_alunos[alunos][4]
        questions[alunos]['acumulateScore'] = wsq_alunos[alunos][6]
        questions[alunos]['answerTime'] = wsq_alunos[alunos][8]
     
print(questions)