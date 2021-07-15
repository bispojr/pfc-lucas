# Comentarios

# overview[1]['playedWith'] = '20'
# finalScore[1]['rank'] = 1
# finalScore[3] = {}
# finalScore[3]['player'] = "Animal"
# questions = { 1: {'score': "", 'statement': "", 'answer': ""}}
# questions[0] = d
# questions[2] = {}
# questions[2][1]['player'] = "Lucas"

# =================================================
#       Estruturas

# overview = { 1: { 'playedOn': "", 'hostedBy': "", 'playedWith': "", 'played': "", 'CorrectAnswers' : "", 'IncorrectAnswers': "", 'Average score': ""}}
# finalScore = { 1: { 'rank': "", 'player': "", 'totalScore': "", 'correctAnswers': "", 'incorrectAnswers': ""}}
# questionSummary = { 1: { 'rank': "", 'player': "", 'totalScore': "", 1: {'score': "", 'statement': "", 'answer': ""}}}
# questions = {'statement': "",'correctAnswers': "", 'playersCorrect': "",'questionDuration': "",'ansOptTriangle': "",'ansOptLosangle': "",'ansOptCircle': "",'ansOptSquare': "",'IsAnswerCorrectTriangle': "",'IsAnswerCorrectLosangle': "",'IsAnswerCorrectCircle': "",'IsAnswerCorrectSquare': "",'NumOfAnsReceivedTriangle': "", 'NumOfAnsReceivedLosangle': "",'NumOfAnsReceivedCircle': "",'NumOfAnsReceivedSquare': "",'TimeToAnsTriangle': "", 'TimeToAnsLosangle': "", 'TimeToAnsCircle': "", 'TimeToAnsSquare': "",
#   1: {
#      'player': "", 'alias': "", 'answerIsCorrect': "", 'statement': "", 'score': "", 'acumulateScore': "", 'answerTime': ""}}

# =============================================

kahoot_18 = [
    r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\2018_kahoot\kahoot_12_09_2018.xlsx',
    r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\2018_kahoot\kahoot_12_09_2018_processos_design_ihc.xlsx',
    r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\2018_kahoot\kahoot_02_10_2018.xlsx',
    r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\2018_kahoot\kahoot_24_10_2018.xlsx',
    r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\2018_kahoot\kahoot_25_10_2018.xlsx',
    r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\2018_kahoot\kahoot_26_10_2018.xlsx',
    r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\2018_kahoot\kahoot_30_10_2018.xlsx',
    r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\2018_kahoot\kahoot_31_10_2018.xlsx',
    r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\2018_kahoot\kahoot_13_11_2018.xlsx',
    r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\2018_kahoot\kahoot_14_11_2018.xlsx']

kahoot_19 = [
    r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\2019_kahoot\results_04_09_2019_conceitos.xlsx',
    r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\2019_kahoot\results_04_09_2019_processos_design.xlsx',
    r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\2019_kahoot\results_01_10_2019_parte_1.xlsx',
    r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\2019_kahoot\results_01_10_2019_parte_2.xlsx',
    r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\2019_kahoot\resultado_09_10_2019.xlsx',
    r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\2019_kahoot\resultado_22_10_2019.xlsx',
    r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\2019_kahoot\resultado_29_10_2019.xlsx',
    r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\2019_kahoot\resultado_30_10_2019.xlsx',
    r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\2019_kahoot\resultado_05_11_2019.xlsx',
    r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\2019_kahoot\resultado_06_11_2019.xlsx',
    r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\2019_kahoot\resultado_19_11_2019.xlsx']

mapping_18 = r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\2018_kahoot\mapeamentos-e-notas-finais.xlsx'
mapping_19 = r'C:\Users\lucas\OneDrive\Documentos\UFG\9º Período - EXTRA\Projeto Final de Curso 2\data\2019_kahoot\notas_com_gamificacao.xlsx'

# ====================== 30/05/2021 =======================


class Lesson:
    def __init__(self):
        self.description = ""
        self.datetime = 0
        self.questions = []
        self.students = []
        self.scores = []


class Student(Lesson):
    def __init__(self):
        self.alias = ""


class Score(Lesson):
    def __init__(self, student, question, answer, isCorrect, isAttend):
        self.student = student
        self.question = question
        self.givenAnswer = answer
        self.isCorrect = isCorrect
        self.isAttend = isAttend

    def get_isCorrect(self):
        return self.isCorrect()

    def set_isCorrect(self, booleano):
        self.isCorrect = booleano

    def get_isAttend(self):
        return False

    def set_isAttend(self, booleano):
        self.isAttend = booleano


class Question(Lesson):
    def __init__(self):
        self.statement = ""
        self.option = []
        self.correctAnswer = 0


class Student():
    def __init__(self):
        self.name = ""
        self.arrayAnswer = []


def mappingStudent(file):
    from openpyxl import load_workbook
    wm = load_workbook(file)
    ws = wm.active

    namesSheets = wm.sheetnames

    mapping = wm[namesSheets[0]]
    nameStudents = []

    for students in range(2, ws.max_row+1):
        nameStudents.append(mapping.cell(row=students, column=1).value)

    return nameStudents


class Parser:
    def __init__(self):
        self.overview = {}
        self.finalScore = {}
        self.questionSummary = {}
        self.questions = {}
        self.nOfSheets = 0
        self.qtdAlunos = 0
        self.maxRowFS = 0
        self.maxRowQS = 0
        self.qs_count = 0
        self.questionsList = 0
        self.question_list = []

    def createDict(self, file):
        from openpyxl import load_workbook
        wb = load_workbook(file)
        ws = wb.active

        self.nOfSheets = len(wb.worksheets)

        # Refatorar - Overview, FInal Score, Question Summary e Questões
        wsOver = wb['Overview']
        wsFinal = wb['Final Scores']
        wsQS = wb['Question Summary']
        wsQues = []

        for q in range(0, (self.nOfSheets - 4)):
            n = str(q + 1)
            # print(q)
            wsQues.append(wb['Question ' + n])

        title = wsOver.cell(row=1, column=1).value
        date = wsOver.cell(row=2, column=2).value
        teacher = wsOver.cell(row=3, column=2).value
        nOfPlayers = wsOver.cell(row=4, column=2).value
        nOfQuestions = wsOver.cell(row=5, column=2).value
        # label_1 = wsOver.cell(row=7, column=1).value
        totalCorrect = wsOver.cell(row=8, column=3).value
        totalIncorrect = wsOver.cell(row=9, column=3).value
        averageScore = wsOver.cell(row=10, column=3).value
        # label_2 = wsOver.cell(row=12, column=1).value
        howFun = wsOver.cell(row=13, column=3).value
        didLearning = []
        didLearning.append(wsOver.cell(row=14, column=3).value)
        didLearning.append(wsOver.cell(row=14, column=5).value)
        doRecommend = []
        doRecommend.append(wsOver.cell(row=15, column=3).value)
        doRecommend.append(wsOver.cell(row=15, column=5).value)
        howFeel = []
        howFeel.append(wsOver.cell(row=16, column=4).value)
        howFeel.append(wsOver.cell(row=16, column=6).value)
        howFeel.append(wsOver.cell(row=16, column=8).value)

        self.overview = {
            'Basic Information': {
                "Lesson Name": title,
                "Played on": date,
                "Hosted by": teacher,
                "Played with": nOfPlayers,
                "Played": nOfQuestions
            },
            "Overall Performance": {
                "Total correct answers (%)": totalCorrect,
                "Total incorrect answers (%)": totalIncorrect,
                "Average score (points)": averageScore
            },
            "Feedback": {
                "How fun was it? (out of 5)": howFun,
                "Did you learn something?": {
                    "Yes": didLearning[0],
                    "No": didLearning[1]
                },
                "Do you recommend it?": {
                    "Yes": doRecommend[0],
                    "No": doRecommend[1]
                },
                "How do you feel?": {
                    "Positive": howFeel[0],
                    "Neutral": howFeel[1],
                    "Negative": howFeel[2]
                }
            }
        }

        # print(overview)
        '''
        print(date)
        print(teacher)
        print(nOfPlayers)
        print(totalCorrect*100, "%")
        print(totalIncorrect*100, "%")
        print(averageScore)
        '''

        # Lista de numeros
        idAl = []
        # Lista de alunos
        student = []
        # Lista de score
        score = []
        # lista de questoes certas
        quesCorrect = []
        # lista de questoes erradas
        quesIncorrect = []

        self.maxRowFS = wsFinal.max_row
        # print(maxRowFS)

        # loop 1: 4 até nOfPlayers+3
        # col=1 and row = i
        for ial in range(4, self.maxRowFS):
            idAl.append(wsFinal.cell(row=ial, column=1).value)

        # loop 2: 4 até nOfPlayers+3
        # col=2 and row = i
        for istd in range(4, self.maxRowFS):
            student.append(wsFinal.cell(row=istd, column=2).value)

        # loop 3: 4 até nOfPlayers+3
        # col=3 and row = i
        for iscor in range(4, self.maxRowFS):
            score.append(wsFinal.cell(row=iscor, column=3).value)

        # loop 4: 4 até nOfPlayers+3
        # col=4 and row = i
        for iqc in range(4, self.maxRowFS):
            quesCorrect.append(wsFinal.cell(row=iqc, column=4).value)

        # loop 5: 4 até nOfPlayers+3
        # col=5 and row = i
        for iqi in range(4, self.maxRowFS):
            quesIncorrect.append(wsFinal.cell(row=iqi, column=5).value)

        self.finalScore = {'Final Scores': {
            'rank': idAl,
            'player': student,
            'totalScore': score,
            'correctAnswers': quesCorrect,
            'incorrectAnswers': quesIncorrect
        }
        }

        # print(finalScore)

        '''
        for iprint in range(0, stopCount-4):
            print(idAl[iprint], "|", student[iprint], "|", score[iprint], "|", quesCorrect[iprint], "|", quesIncorrect[iprint])
        '''

        qs_idAl = []
        qs_student = []
        qs_totalScore = []

        numQuestions = self.nOfSheets - 4
        self.qs_count = numQuestions
        self.maxRowQS = wsQS.max_row
        qs_score = [[] for _ in range(self.qs_count)]
        qs_statement = [[] for _s_ in range(self.qs_count)]
        qs_answer = [[] for __ in range(self.qs_count)]

        for i in range(4, self.maxRowQS):
            qs_idAl.append(wsQS.cell(row=i, column=1).value)
            qs_student.append(wsQS.cell(row=i, column=2).value)
            qs_totalScore.append(wsQS.cell(row=i, column=3).value)
            for j in range(0, self.qs_count):
                col = 4 + 2 * j
                col2 = 5 + 2 * j
                qs_score[j].append(wsQS.cell(row=i, column=col).value)
                qs_statement[j].append(wsQS.cell(row=3, column=col2).value)
                qs_answer[j].append(wsQS.cell(row=i, column=col2).value)

        self.questionSummary = {'Question Summary': {
            'rank': qs_idAl,
            'player': qs_student,
            'totalScore': qs_totalScore
        },
            1: {
                'score': qs_score,
                'statement': qs_statement,
                'answer': qs_answer
            }
        }

        # print(questionSummary)
        '''
        for iprint in range(0, stopCount-4):
            print(qs_idAl[iprint], "|", qs_student[iprint], "|", qs_totalScore[iprint], "|")
            for qprint in range(0, qs_count-1):
                print(qs_score[qprint][iprint], "|", qs_answer[qprint][iprint], "|")
            print("")
        '''

        # self.question_list = []
        self.qtdAlunos = self.maxRowQS - 5
        for quest in range(0, numQuestions):

            wsq_numq = []
            wsq_statement = []
            wsq_correctAnswers = []
            wsq_playersCorrect = []
            wsq_questionDuration = []
            wsq_ansOptTriangle = []
            wsq_ansOptLosangle = []
            wsq_ansOptCircle = []
            wsq_ansOptSquare = []
            wsq_IsAnswerCorrectTriangle = []
            wsq_IsAnswerCorrectLosangle = []
            wsq_IsAnswerCorrectCircle = []
            wsq_IsAnswerCorrectSquare = []
            wsq_NumOfAnsReceivedTriangle = []
            wsq_NumOfAnsReceivedLosangle = []
            wsq_NumOfAnsReceivedCircle = []
            wsq_NumOfAnsReceivedSquare = []
            wsq_TimeToAnsTriangle = []
            wsq_TimeToAnsLosangle = []
            wsq_TimeToAnsCircle = []
            wsq_TimeToAnsSquare = []

            wsq_alunos = [[] for ____ in range(self.qtdAlunos)]

            for k in range(0, self.qtdAlunos):
                for l in range(1, 11):
                    wsq_alunos[k].append(wsQues[quest].cell(row=k + 15, column=l).value)

            wsq_numq.append(wsQues[quest].cell(row=2, column=1).value)
            wsq_statement.append(wsQues[quest].cell(row=2, column=2).value)
            wsq_correctAnswers.append(wsQues[quest].cell(row=3, column=3).value)
            wsq_playersCorrect.append(wsQues[quest].cell(row=4, column=3).value)
            wsq_questionDuration.append(wsQues[quest].cell(row=5, column=3).value)
            wsq_ansOptTriangle.append(wsQues[quest].cell(row=8, column=4).value)
            wsq_ansOptLosangle.append(wsQues[quest].cell(row=8, column=6).value)
            wsq_ansOptCircle.append(wsQues[quest].cell(row=8, column=8).value)
            wsq_ansOptSquare.append(wsQues[quest].cell(row=8, column=10).value)

            if (wsQues[quest].cell(row=9, column=3).value) == "✔︎":
                wsq_IsAnswerCorrectTriangle.append(True)
            else:
                wsq_IsAnswerCorrectTriangle.append(False)

            if (wsQues[quest].cell(row=9, column=5).value) == "✔︎":
                wsq_IsAnswerCorrectLosangle.append(True)
            else:
                wsq_IsAnswerCorrectLosangle.append(False)

            if (wsQues[quest].cell(row=9, column=7).value) == "✔︎":
                wsq_IsAnswerCorrectCircle.append(True)
            else:
                wsq_IsAnswerCorrectCircle.append(False)

            if (wsQues[quest].cell(row=9, column=9).value) == "✔︎":
                wsq_IsAnswerCorrectSquare.append(True)
            else:
                wsq_IsAnswerCorrectSquare.append(False)

            wsq_NumOfAnsReceivedTriangle.append(wsQues[quest].cell(row=10, column=3).value)
            wsq_NumOfAnsReceivedLosangle.append(wsQues[quest].cell(row=10, column=5).value)
            wsq_NumOfAnsReceivedCircle.append(wsQues[quest].cell(row=10, column=7).value)
            wsq_NumOfAnsReceivedSquare.append(wsQues[quest].cell(row=10, column=9).value)
            wsq_TimeToAnsTriangle.append(wsQues[quest].cell(row=11, column=3).value)
            wsq_TimeToAnsLosangle.append(wsQues[quest].cell(row=11, column=5).value)
            wsq_TimeToAnsCircle.append(wsQues[quest].cell(row=11, column=7).value)
            wsq_TimeToAnsSquare.append(wsQues[quest].cell(row=11, column=9).value)

            if wsq_alunos[0][2] == "✔︎":
                wsq_alunos[0][2] = True
            else:
                wsq_alunos[0][2] = False

            self.questions = {'Basic informations': {
                'numOfQuestion': wsq_numq,
                'statement': wsq_statement,
                'correctAnswers': wsq_correctAnswers,
                'playersCorrect': wsq_playersCorrect,
                'questionDuration': wsq_questionDuration
            },
                "Answare Summary": {
                    'ansOptTriangle': wsq_ansOptTriangle,
                    'ansOptLosangle': wsq_ansOptLosangle,
                    'ansOptCircle': wsq_ansOptCircle,
                    'ansOptSquare': wsq_ansOptSquare,
                    'IsAnswerCorrectTriangle': wsq_IsAnswerCorrectTriangle,
                    'IsAnswerCorrectLosangle': wsq_IsAnswerCorrectLosangle,
                    'IsAnswerCorrectCircle': wsq_IsAnswerCorrectCircle,
                    'IsAnswerCorrectSquare': wsq_IsAnswerCorrectSquare,
                    'NumOfAnsReceivedTriangle': wsq_NumOfAnsReceivedTriangle,
                    'NumOfAnsReceivedLosangle': wsq_NumOfAnsReceivedLosangle,
                    'NumOfAnsReceivedCircle': wsq_NumOfAnsReceivedCircle,
                    'NumOfAnsReceivedSquare': wsq_NumOfAnsReceivedSquare,
                    'TimeToAnsTriangle': wsq_TimeToAnsTriangle,
                    'TimeToAnsLosangle': wsq_TimeToAnsLosangle,
                    'TimeToAnsCircle': wsq_TimeToAnsCircle,
                    'TimeToAnsSquare': wsq_TimeToAnsSquare,
                },
                'Answer Details': {
                    1: {
                        'player': wsq_alunos[0][0],
                        'alias': wsq_alunos[0][1],
                        'answerIsCorrect': wsq_alunos[0][2],
                        'statement': wsq_alunos[0][3],
                        'score': wsq_alunos[0][4],
                        'acumulateScore': wsq_alunos[0][6],
                        'answerTime': wsq_alunos[0][8]
                    }
                }
            }
            for alunos in range(1, self.qtdAlunos + 1):
                self.questions[alunos] = {}
                _alunos = alunos - 1
                self.questions[alunos]['player'] = wsq_alunos[_alunos][0]
                self.questions[alunos]['alias'] = wsq_alunos[_alunos][1]
                if wsq_alunos[_alunos][2] == "✘":
                    wsq_alunos[_alunos][2] = False
                else:
                    wsq_alunos[_alunos][2] = True
                self.questions[alunos]['answerIsCorrect'] = wsq_alunos[_alunos][2]
                self.questions[alunos]['statement'] = wsq_alunos[_alunos][3]
                self.questions[alunos]['score'] = wsq_alunos[_alunos][4]
                self.questions[alunos]['acumulateScore'] = wsq_alunos[_alunos][6]
                self.questions[alunos]['answerTime'] = wsq_alunos[_alunos][8]

            self.question_list.append(self.questions)
            # print(questions)

        return wb

    def toJSON(self, wb):
        # ======================  20/05/2021 ==================================

        nomes = wb.sheetnames

        # ==================  OVERVIEW ==================
        txt_overview = ('{"' + nomes[0] + '": {' + "\n"
                        + "\t" + '"' + self.overview['Basic Information']["Lesson Name"] + '": {' + "\n"
                        + "\t\t" + '"' + "Played on" + '": "' + str(
                    self.overview['Basic Information']["Played on"]) + '", ' + "\n"
                        + "\t\t" + '"' + "Hosted by" + '": "' + str(
                    self.overview['Basic Information']["Hosted by"]) + '", ' + "\n"
                        + "\t\t" + '"' + "Played with" + '": "' + str(
                    self.overview['Basic Information']["Played with"]) + '", ' + "\n"
                        + "\t\t" + '"' + "Played" + '": "' + str(
                    self.overview['Basic Information']["Played"]) + '"' + "\n"
                        + "\t" + "}, " + "\n"
                        + "\t" + '"' + "Overall Performance" + '": {' + "\n"
                        + "\t\t" + '"' + "Total correct answers (%)" + '": "' + str(
                    self.overview['Overall Performance']["Total correct answers (%)"]) + '", ' + "\n"
                        + "\t\t" + '"' + "Total incorrect answers (%)" + '": "' + str(
                    self.overview['Overall Performance']["Total incorrect answers (%)"]) + '", ' + "\n"
                        + "\t\t" + '"' + "Average score (points)" + '": "' + str(
                    self.overview['Overall Performance']["Average score (points)"]) + '"' + "\n"
                        + "\t" + "}," + "\n"
                        + "\t" + '"' + "Feedback" + '": {' + "\n"
                        + "\t\t" + '"' + "How fun was it? (out of 5)" + '": "' + str(
                    self.overview['Feedback']["How fun was it? (out of 5)"]) + '", ' + "\n"
                        + "\t\t" + '"' + "Did you learn something?" + '": {' + "\n"
                        + "\t\t\t" + '"' + "Yes" + '": "' + str(
                    self.overview['Feedback']["Did you learn something?"]["Yes"]) + '", ' + "\n"
                        + "\t\t\t" + '"' + "No" + '": "' + str(
                    self.overview['Feedback']["Did you learn something?"]["No"]) + '"' + "\n"
                        + "\t\t" + "}," + "\n"
                        + "\t\t" + '"' + "Do you recommend it?" + '": {' + "\n"
                        + "\t\t\t" + '"' + "Yes" + '": "' + str(
                    self.overview['Feedback']["Do you recommend it?"]["Yes"]) + '", ' + "\n"
                        + "\t\t\t" + '"' + "No" + '": "' + str(
                    self.overview['Feedback']["Do you recommend it?"]["No"]) + '"' + "\n"
                        + "\t\t" + "}," + "\n"
                        + "\t\t" + '"' + "How do you feel?" + '": {' + "\n"
                        + "\t\t\t" + '"' + "Positive" + '": "' + str(
                    self.overview['Feedback']["How do you feel?"]["Positive"]) + '", ' + "\n"
                        + "\t\t\t" + '"' + "Neutral" + '": "' + str(
                    self.overview['Feedback']["How do you feel?"]["Neutral"]) + '", ' + "\n"
                        + "\t\t\t" + '"' + "Negative" + '": "' + str(
                    self.overview['Feedback']["How do you feel?"]["Negative"]) + '"' "\n"
                        + "\t\t" + "}" + "\n"
                        + "\t" + "}\n" + "}, \n")

        # ==================  FINAL SCORE ==================

        txt_final = '"' + nomes[1] + '": {' + "\n" + "\t" + '"' + self.overview['Basic Information'][
            "Lesson Name"] + '": \n' + "\t" + "["
        txt_final_students = []

        for t in range(0, self.maxRowFS - 5):
            txt_final_students.append("\t\t{\n"
                                      + "\t\t\t" + '"' + "Rank" + '": "' + str(
                self.finalScore['Final Scores']['rank'][t]) + '", ' + "\n"
                                      + "\t\t\t" + '"' + 'Player' + '": "' + str(
                self.finalScore['Final Scores']['player'][t]) + '", ' + "\n"
                                      + "\t\t\t" + '"' + 'Total Score' + '": "' + str(
                self.finalScore['Final Scores']['totalScore'][t]) + '", ' + "\n"
                                      + "\t\t\t" + '"' + 'Correct Answers' + '": "' + str(
                self.finalScore['Final Scores']['correctAnswers'][t]) + '", ' + "\n"
                                      + "\t\t\t" + '"' + 'Incorrect Answers' + '": "' + str(
                self.finalScore['Final Scores']['incorrectAnswers'][t]) + '"' + "\n")
            if t == self.maxRowFS - 6:
                txt_final_students[t] = txt_final_students[t] + "\t\t" + "}" + "\n"
            else:
                txt_final_students[t] = txt_final_students[t] + "\t\t" + "}," + "\n"

        txt_final_students[self.maxRowFS - 6] = txt_final_students[self.maxRowFS - 6] + "\t]\n},"
        # ==================  QUESTION SUMMARY ==================

        txt_questionSummary_students = []
        txt_questionSummary_students_questions = []
        concate = []

        txt_questionSummary = '"' + nomes[2] + '": {' + "\n" + "\t" + '"' + self.overview['Basic Information'][
            "Lesson Name"] + '": {' \
                              + "\n" + "\t\t" + '"' + nomes[2] + '": ' + "\n" + "\t\t" + "["

        n = 1

        # questionSummary["Alunos"]['score'][0][1]
        # questionSummary["Alunos"][_CAMPO_][_QUESTÃO_][_ALUNO_]

        for q in range(0, self.maxRowQS - 5):
            txt_questionSummary_students.append("\t\t\t" + "{\n"
                                                + "\t\t\t\t" + '"' + 'Rank' + '": "' + str(
                self.questionSummary['Question Summary']['rank'][q]) + '", ' + "\n"
                                                + "\t\t\t\t" + '"' + 'Player' + '": "' + str(
                self.questionSummary['Question Summary']['player'][q]) + '", ' + "\n"
                                                + "\t\t\t\t" + '"' + 'Total Score' + '": "' + str(
                self.questionSummary['Question Summary']['totalScore'][q]) + '", ' + "\n")
            concatenar = txt_questionSummary_students[q]
            # print(concatenar)
            for s in range(0, self.qs_count):
                txt_questionSummary_students_questions.append("\t\t\t\t" + '"' + 'Q' + str(s + 1) + '": {' + "\n"
                                                              + "\t\t\t\t\t\t" + '"' + "Score" + '": "' + str(
                    self.questionSummary[1]['score'][s][q]) + '", ' + "\n"
                                                              + "\t\t\t\t\t\t" + '"' + "Statement" + '": "' + str(
                    self.questionSummary[1]['statement'][s][q]) + '", ' + "\n"
                                                              + "\t\t\t\t\t\t" + '"' + "Answer" + '": "' + str(
                    self.questionSummary[1]['answer'][s][q]) + '"' + "\n")
                if s == self.qs_count - 1:
                    txt_questionSummary_students_questions[s] = txt_questionSummary_students_questions[s] + (
                            "\t\t\t\t" + "}" + "\n")
                else:
                    txt_questionSummary_students_questions[s] = txt_questionSummary_students_questions[
                                                                    s] + "\t\t\t\t" + "}, " + "\n"

                concatenar = concatenar + txt_questionSummary_students_questions[s]
            txt_questionSummary_students_questions = []
            if q == self.maxRowQS - 6:
                concatenar = concatenar + "\t\t\t" + "}" + "\n"
            else:
                concatenar = concatenar + "\t\t\t" + "}, " + "\n"
            concate.append(concatenar)

        ind = self.maxRowQS - 6
        concate[ind] = concate[ind] + "\t\t]\n" + "\t" + "}\n" + "},\n"
        # ==================  QUESTIONS ==================

        txt_question = []
        txt_question_alunos = []
        concate_question = []

        for qn in range(0, self.qs_count):
            txt_question.append('"' + nomes[qn + 3] + '": {' + "\n" + "\t"
                                + '"' + self.overview['Basic Information']["Lesson Name"] + '": {' + "\n"
                                + "\t\t" + '"' + self.question_list[qn]['Basic informations']['numOfQuestion'][
                                    0] + '": {' + "\n"
                                + "\t\t\t" + '"' + "Statement" + '": "' +
                                self.question_list[qn]['Basic informations']['statement'][
                                    0] + '", ' + "\n"
                                + "\t\t\t" + '"' + "Correct Answer" + '": "' + str(
                self.question_list[qn]['Basic informations']['correctAnswers'][0]) + '", ' + "\n"
                                + "\t\t\t" + '"' + "Players Correct" + '": "' + str(
                self.question_list[qn]['Basic informations']['playersCorrect'][0]) + '", ' + "\n"
                                + "\t\t\t" + '"' + "Question Duration" + '": "' + str(
                self.question_list[qn]['Basic informations']['questionDuration'][0]) + '"' + "\n"
                                + "\t\t" + "},\n"
                                + "\t\t" + '"' + "Answare Summary" + '": {' + "\n"
                                + "\t\t\t" + '"' + "Answer options" + '": {' + "\n"
                                + "\t\t\t\t\t" + '"' + "Triangle" + '": "' + str(
                self.question_list[qn]['Answare Summary']['ansOptTriangle'][0]) + '", ' + "\n"
                                + "\t\t\t\t\t" + '"' + "Losangle" + '": "' + str(
                self.question_list[qn]['Answare Summary']['ansOptLosangle'][0]) + '", ' + "\n"
                                + "\t\t\t\t\t" + '"' + "Circle" + '": "' + str(
                self.question_list[qn]['Answare Summary']['ansOptCircle'][0]) + '", ' + "\n"
                                + "\t\t\t\t\t" + '"' + "Square" + '": "' + str(
                self.question_list[qn]['Answare Summary']['ansOptSquare'][0]) + '"' + "\n"
                                + "\t\t\t" + "},\n"
                                + "\t\t\t" + '"' + "Is answer correct?" + '": {' + "\n"
                                + "\t\t\t\t\t" + '"' + "Triangle" + '": "' + str(
                self.question_list[qn]['Answare Summary']['IsAnswerCorrectTriangle'][0]) + '", ' + "\n"
                                + "\t\t\t\t\t" + '"' + "Losangle" + '": "' + str(
                self.question_list[qn]['Answare Summary']['IsAnswerCorrectLosangle'][0]) + '", ' + "\n"
                                + "\t\t\t\t\t" + '"' + "Circle" + '": "' + str(
                self.question_list[qn]['Answare Summary']['IsAnswerCorrectCircle'][0]) + '", ' + "\n"
                                + "\t\t\t\t\t" + '"' + "Square" + '": "' + str(
                self.question_list[qn]['Answare Summary']['IsAnswerCorrectSquare'][0]) + '"' + "\n"
                                + "\t\t\t" + "},\n"
                                + "\t\t\t" + '"' + "Number of answers received" + '": {' + "\n"
                                + "\t\t\t\t\t" + '"' + "Triangle" + '": "' + str(
                self.question_list[qn]['Answare Summary']['NumOfAnsReceivedTriangle'][0]) + '", ' + "\n"
                                + "\t\t\t\t\t" + '"' + "Losangle" + '": "' + str(
                self.question_list[qn]['Answare Summary']['NumOfAnsReceivedLosangle'][0]) + '", ' + "\n"
                                + "\t\t\t\t\t" + '"' + "Circle" + '": "' + str(
                self.question_list[qn]['Answare Summary']['NumOfAnsReceivedCircle'][0]) + '", ' + "\n"
                                + "\t\t\t\t\t" + '"' + "Square" + '": "' + str(
                self.question_list[qn]['Answare Summary']['NumOfAnsReceivedSquare'][0]) + '"' + "\n"
                                + "\t\t\t" + "},\n"
                                + "\t\t\t" + '"' + "Average time taken to answer (seconds)" + '": {' + "\n"
                                + "\t\t\t\t\t" + '"' + "Triangle" + '": "' + str(
                self.question_list[qn]['Answare Summary']['TimeToAnsTriangle'][0]) + '", ' + "\n"
                                + "\t\t\t\t\t" + '"' + "Losangle" + '": "' + str(
                self.question_list[qn]['Answare Summary']['TimeToAnsLosangle'][0]) + '", ' + "\n"
                                + "\t\t\t\t\t" + '"' + "Circle" + '": "' + str(
                self.question_list[qn]['Answare Summary']['TimeToAnsCircle'][0]) + '", ' + "\n"
                                + "\t\t\t\t\t" + '"' + "Square" + '": "' + str(
                self.question_list[qn]['Answare Summary']['TimeToAnsSquare'][0]) + '"' + "\n"
                                + "\t\t\t" + "}\n"
                                + "\t\t" + "},\n"
                                + "\t\t" + '"' + "Answer Details" + '": ' + "\n" + "\t\t[\n")
            concatenar_Question = txt_question[qn]
            for qnn in range(0, self.qtdAlunos):
                qn_n = qnn + 1
                txt_question_alunos.append("\t\t\t" + "{\n"
                                           + "\t\t\t\t" + '"' + "Player" + '": "' + str(
                    self.question_list[qn][qn_n]['player']) + '", ' + "\n"
                                           + "\t\t\t\t" + '"' + "Alias" + '": "' + str(
                    self.question_list[qn][qn_n]['alias']) + '", ' + "\n"
                                           + "\t\t\t\t" + '"' + "Answer" + '": {' + "\n"
                                           + "\t\t\t\t\t\t" + '"' + "Correct?" + '": "' + str(
                    self.question_list[qn][qn_n]['answerIsCorrect']) + '", ' + "\n"
                                           + "\t\t\t\t\t\t" + '"' + "Statement" + '": "' + str(
                    self.question_list[qn][qn_n]['statement']) + '" ' + "\n"
                                           + "\t\t\t\t" + "}," + "\n"
                                           + "\t\t\t\t" + '"' + "Score (points)" + '": "' + str(
                    self.question_list[qn][qn_n]['score']) + '", ' + "\n"
                                           + "\t\t\t\t" + '"' + "Current Total Score (points)" + '": "' + str(
                    self.question_list[qn][qn_n]['acumulateScore']) + '", ' + "\n"
                                           + "\t\t\t\t" + '"' + "Answer time (seconds)" + '": "' + str(
                    self.question_list[qn][qn_n]['answerTime']) + '"' + "\n")
                if qnn == self.qtdAlunos - 1:
                    txt_question_alunos[qnn] = txt_question_alunos[qnn] + "\t\t\t" + "}\n"
                else:
                    txt_question_alunos[qnn] = txt_question_alunos[qnn] + "\t\t\t" + "},\n"
                concatenar_Question = concatenar_Question + txt_question_alunos[qnn]
            txt_question_alunos = []
            concatenar_Question = concatenar_Question + "\t\t" + "]\n" + "\t" + "}\n"
            if qn == self.qs_count - 1:
                concatenar_Question = concatenar_Question + "}\n}"
            else:
                concatenar_Question = concatenar_Question + "},\n"
            concate_question.append(concatenar_Question)

        # print(txt_overview)
        # print(txt_final)
        # print(*txt_final_students)
        # print(txt_questionSummary)
        # print(*concate)
        # print(question_list[1])
        # print(*txt_question)
        # print(*concate_question)

        arquivo = open('C:/Users/lucas/OneDrive/Documentos/UFG/9º Período - EXTRA/Projeto Final de Curso 2/data/2018_kahoot_JSON/' + self.overview['Basic Information']["Lesson Name"] + ".json", 'w', encoding="utf-8")
        arquivo.write(txt_overview)
        arquivo.write("\n")
        arquivo.write(txt_final)
        arquivo.write(''.join(txt_final_students))
        arquivo.write("\n")
        arquivo.write(''.join(txt_questionSummary))
        arquivo.write(''.join(concate))
        arquivo.write("\n")
        arquivo.write(''.join(concate_question))
        arquivo.close()

    def toLesson(self):
        thisLesson = Lesson()

        thisLesson.description = self.overview['Basic Information']["Lesson Name"]
        thisLesson.datetime = self.overview['Basic Information']["Played on"]

        students = []
        questions = []
        score = []

        for k in range(0, self.qs_count):
            questions.append(Question())
            questions[k].statement = self.question_list[k]['Basic informations']['statement'][0]
            questions[k].option.append(self.question_list[k]['Answare Summary']['ansOptTriangle'][0])
            questions[k].option.append(self.question_list[k]['Answare Summary']['ansOptLosangle'][0])
            questions[k].option.append(self.question_list[k]['Answare Summary']['ansOptCircle'][0])
            questions[k].option.append(self.question_list[k]['Answare Summary']['ansOptSquare'][0])
            if str(self.question_list[k]['Answare Summary']['IsAnswerCorrectTriangle'][0]) == "True":
                questions[k].correctAnswer = 1
            if str(self.question_list[k]['Answare Summary']['IsAnswerCorrectLosangle'][0]) == "True":
                questions[k].correctAnswer = 2
            if str(self.question_list[k]['Answare Summary']['IsAnswerCorrectCircle'][0]) == "True":
                questions[k].correctAnswer = 3
            if str(self.question_list[k]['Answare Summary']['IsAnswerCorrectSquare'][0]) == "True":
                questions[k].correctAnswer = 4

        for i in range(0, self.qtdAlunos):
            j = i + 1
            students.append(Student())
            students[i].alias = self.questions[j]['alias']
            thisLesson.students.append(students[i])

        for scA in range(0, self.qtdAlunos):
            for scQ in range(0, self.qs_count):
                ans = self.questionSummary[1]['answer'][scQ][scA]
                correct = self.question_list[1][2]['answerIsCorrect']
                if ans == "" or ans == " ":
                    attend = False
                else:
                    attend = True
                score.append(Score(students[scA], questions[scQ], ans, correct, attend))

        # print(questions[0].statement + " | " + students[0].description + " | " + str(questions[0].correctAnswer))
        thisLesson.students = students
        thisLesson.questions = questions
        thisLesson.score = score

        return thisLesson

    def toArray(self, lesson):
        n = self.qs_count
        c = 0
        for i in range(0, len(lesson.scores)):
            c = c + 1
            if lesson.scores[i].isAttend == False:
                self.arrayAnswers.append(0)
            elif lesson.scores[i].isCorrect == True:
                self.arrayAnswers.append(1)
            elif lesson.scores[i].isCorrect == False:
                self.arrayAnswers.append(-1)
            if c == n:
                self.arrayStudents.append(self.arrayAnswers)
                c = 0
                self.arrayAnswers = []

read_2018 = []
work_2018 = []
lesson_2018 = []
arrays_2018 = []

for ler18 in range(0, 10):
    read_2018.append(Parser())
    work_2018.append(read_2018[ler18].createDict(kahoot_18[ler18], 2018))
    read_2018[ler18].toJSON(work_2018[ler18])
    lesson_2018.append(read_2018[ler18].toLesson())
    read_2018[ler18].toArray(lesson_2018[ler18])

read_2019 = []
work_2019 = []
lesson_2019 = []

for ler19 in range(0, 11):
    read_2019.append(Parser())
    work_2019.append(read_2019[ler19].createDict(kahoot_19[ler19], 2019))
    read_2019[ler19].toJSON(work_2019[ler19])
    lesson_2019.append(read_2019[ler19].toLesson())

mapp18 = mappingStudent(mapping_18)
print(mapp18)
mapp19 = mappingStudent(mapping_19)
print(mapp19)

student = []
total_Alunos18 = len(mapp18)
total_Alunos19 = len(mapp19)

# for aulas in range(0, len(kahoot_18)):
#     user = 0
#     step = read_2018[aulas].question_count #quantidade de questoes
#     print(read_2018[aulas].overview['Basic Information']["Played on"])
#     for vetor_aula in range(0, len(read_2018[aulas].arrayStudents)):
#         print(lesson_2018[aulas].scores[user].student.alias + " | " + str(read_2018[aulas].arrayStudents[vetor_aula]))
#         user = user + step #pula a quantidade de questoes
#     print("========================")

# for aluno in student:
#     print(aluno.name + "\n" + str(aluno.arrayAnswer))

# for b in range(0, len(kahoot_19)):
#     read_2019[b].toArray(lesson_2019[b])
#     t = 0
#     z = read_2019[b].qs_count
#     print(read_2019[b].overview['Basic Information']["Played on"])
#     for x in range(0, len(read_2019[b].arrayStudents)):
#         print(lesson_2019[b].scores[t].student.alias + " | " + str(read_2019[b].arrayStudents[x]))
#         t = t + z
#     print("========================")

# print(read_2018[0].qs_count)
# n = read_2018[0].qs_count
# c = 0
# print(read_2018[0].question_list[0][3]['answerIsCorrect'])
# print(read_2018[0].question_list[0][3]['alias'])
# print(read_2018[0].question_list[0]['Basic informations']['numOfQuestion'][0])
# for i in range(0, len(lesson_2018[0].scores)):
#     c = c + 1
#     print(lesson_2018[0].scores[i].student.alias + " | " + str(lesson_2018[0].scores[i].isAttend) + " | " + str(lesson_2018[0].scores[i].isCorrect))
#     if c == n:
#         print("=========")
#         c = 0